# compare_files.py
import openpyxl
from importlib import import_module
from select_files import select_files

# Lazy imports
display_results = import_module("display_results")
# # compare_files.py
# import os
# import pandas as pd
# import openpyxl
# from importlib import import_module
# from display_results import display_results
# from select_files import select_files

# # Lazy imports
# display_results = import_module("display_results")
# select_files = import_module("select_files")

def compare(arquivos):
    # """
    # Compara dois arquivos Excel.

    # Args:
    #     arquivos: Lista de caminhos para os arquivos Excel.

    # Returns:
    #     Lista de listas, onde cada lista representa uma linha de dados dos arquivos comparados.
    # """

    # Abre os arquivos Excel
    arquivo_1 = openpyxl.load_workbook(arquivos[0])
    arquivo_2 = openpyxl.load_workbook(arquivos[1])

    # Obtém as planilhas dos arquivos
    planilha_1 = arquivo_1.active
    planilha_2 = arquivo_2.active

    # Obtém as colunas dos arquivos
    colunas_1 = list(planilha_1.columns)
    colunas_2 = list(planilha_2.columns)

    # Compara as linhas dos arquivos
    resultados = []
    for linha in range(planilha_1.max_row):
        linha_1 = [planilha_1[coluna][linha].value for coluna in colunas_1]
        linha_2 = [planilha_2[coluna][linha].value for coluna in colunas_2]
        resultados.append(linha_1 == linha_2)
    colunas = colunas_1

    return resultados, colunas  # Retornando também a lista de colunas


def main():
    # Seleciona os arquivos a serem comparados
    caminhos = select_files()

    # Compara os arquivos
    resultados, colunas = compare(caminhos)

    # Exibe os resultados
    for coluna in range(len(resultados[0])):
        display_results.display_results(resultados, coluna)


if __name__ == "__main__":
    main()
# def compare(arquivos):
#     # """
#     # Compara dois arquivos Excel.

#     # Args:
#     #     arquivos: Lista de caminhos para os arquivos Excel.

#     # Returns:
#     #     Lista de listas, onde cada lista representa uma linha de dados dos arquivos comparados.
#     # """

#     # Abre os arquivos Excel
#     arquivo_1 = openpyxl.load_workbook(arquivos[0])
#     arquivo_2 = openpyxl.load_workbook(arquivos[1])

#     # Obtém as planilhas dos arquivos
#     planilha_1 = arquivo_1.active
#     planilha_2 = arquivo_2.active

#     # Obtém as colunas dos arquivos
#     colunas_1 = list(planilha_1.columns)
#     colunas_2 = list(planilha_2.columns)

#     # Compara as linhas dos arquivos
#     resultados = []
#     for linha in range(planilha_1.max_row):
#         linha_1 = [planilha_1[coluna][linha].value for coluna in colunas_1]
#         linha_2 = [planilha_2[coluna][linha].value for coluna in colunas_2]
#         resultados.append(linha_1 == linha_2)
#     colunas = colunas_1

#     return resultados


# def main():
#     # Seleciona os arquivos a serem comparados
#     caminhos = select_files()

#     # Compara os arquivos
#     resultados = compare(caminhos)

#     # Exibe os resultados
#     for coluna in range(len(resultados[0])):
#         display_results(resultados, coluna)


# if __name__ == "__main__":
#     main()


# def compare_files(file_paths):
#     # """
#     # Compara dois arquivos Excel.

#     # Args:
#     #     file_paths: Lista de caminhos dos arquivos Excel.

#     # Returns:
#     #     Lista de resultados.
#     # """

#     if len(file_paths) != 2:
#         raise ValueError("Selecione dois arquivos Excel.")

#     df1 = pd.read_excel(file_paths[0])
#     df2 = pd.read_excel(file_paths[1])

#     colunas_comparacao = ["NR_NOTA", "VALOR"]
#     resultados = []
#     for coluna in colunas_comparacao:
#         if coluna not in df1.columns or coluna not in df2.columns:
#             resultados.append(f"Coluna {coluna} não existe em nenhum dos arquivos.")
#         else:
#             valores1 = df1[coluna].tolist()
#             valores2 = df2[coluna].tolist()
#             if valores1 != valores2:
#                 resultados.append(f"Diferença na coluna {coluna}: {valores1} != {valores2}")

#     return resultados
